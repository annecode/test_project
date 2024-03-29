#!/usr/bin/env python
# -*- coding: utf-8 -*-
# author: anne-home
# datetime: 2020-12-14 23:11
# filename: test_project/parse_excel

import os
from openpyxl import load_workbook


class ParseExcel:
    def __init__(self, excelPath, sheetName):
        self.wb = load_workbook(excelPath)  # 读取excel文件，创建对象
        self.sheet = self.wb[sheetName]
        self.maxRowNum = self.sheet.max_row

    def get_datas_from_sheet(self):
        dataList = []
        for line in self.sheet.rows:  # line是元组，逐行读取
            lineList = (line[0].value, line[1].value)
            dataList.append(lineList)
        return dataList[1:]


if __name__ == '__main__':
    excelPath = os.path.join(os.path.dirname(__file__), '../test_data/data_excel.xlsx')
    sheetName = 'testcase1'
    excel = ParseExcel(excelPath, sheetName)
    result = excel.get_datas_from_sheet()
    print(*result)
